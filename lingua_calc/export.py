"""Workbook export of a report (issue #3, now xlsx).

Exporting is a per-run action: the UI hangs it off the run in the history table,
so a stored run can be pulled without first rendering it and without disturbing
whatever is already on screen. It is built here, server-side, from the same
``MultiFileReport`` the report route returns — the displayed grain, re-derived
from the store for free (no provider call).

It lives on the server rather than in the browser because xlsx is a zip of XML
parts, not a line-oriented text format: the CSV this replaces was cheap to
assemble in `app.js`, a workbook is not. What the move buys, beyond not
hand-rolling a zip writer in the UI, is that the grains below are testable in
the same suite as the statistics they carry. The cost is unchanged from the CSV:
a report reachable without a stored run behind it (LINGUA_PERSIST_RUNS=false
hides history entirely) has no export path; re-enable persistence to export.

Two sheets, because the workbook answers two different questions.

``Rows`` is the flat grain, one row per chapter × lemma × parse — the rows the
chapter table shows, with the file and chapter they belong to spliced in so the
whole corpus lands in one sheet that pivots.

``New lemmas by chapter`` is the vocabulary question on its own: only the rows
where a lemma appears for the first time in the corpus, collapsed to one row per
lemma. Derivable from the first sheet (filter ``first_occ_lemma``, then dedupe
the parse rows down to the lemma), but that is a pivot the reader would have to
rebuild every time, and "what does this chapter introduce" is the question the
whole cumulative apparatus exists to answer.

Order in both is report order — first appearance within the chapter — not
whatever the on-screen sort happens to be. The sort is a reading aid, and a
spreadsheet re-sorts anyway.
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from lingua_calc.models import MultiFileReport

# `chapter_index` and the four `*_chapter` columns are the model's 0-based
# corpus-wide indexes, left raw so they compare against each other: a row is a
# lemma's first appearance exactly when lemma_first_chapter == chapter_index.
# The booleans alongside them are that comparison already done, because "is this
# the first time?" is the question the author actually asks.
ROW_HEADER = [
    "file",
    "chapter_index",
    "chapter_id",
    "chapter_title",
    "type",
    "lemma",
    "form",
    "parse",
    "lemma_occ",
    "parse_occ",
    "form_occ",
    "first_occ_lemma",
    "first_occ_parse",
    "last_occ_lemma",
    "last_occ_parse",
    "lemma_first_chapter",
    "lemma_last_chapter",
    "parse_first_chapter",
    "parse_last_chapter",
]

# `lemma_occ` is the lemma's count in the chapter that introduces it, so the
# pair (lemma_occ, lemma_last_chapter) says both how hard the chapter drills a
# new word and whether it ever comes back. `forms` and `parses` are the shapes
# it is met in, joined rather than exploded: exploding them would put the lemma
# on several rows again, which is exactly what this sheet collapses.
NEW_LEMMA_HEADER = [
    "file",
    "chapter_index",
    "chapter_id",
    "chapter_title",
    "type",
    "lemma",
    "lemma_occ",
    "form_count",
    "forms",
    "parse_count",
    "parses",
    "lemma_last_chapter",
    "chapter_span",
]

ROW_SHEET = "Rows"
NEW_LEMMA_SHEET = "New lemmas by chapter"

MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Separator for the joined form/parse columns. Not a comma: a spreadsheet cell
# full of commas reads as a CSV that failed to split, and these are Greek words
# that may carry their own punctuation.
JOIN = " · "


def export_rows(report: MultiFileReport) -> list[list]:
    """The flat chapter × lemma × parse grain."""
    rows: list[list] = []
    for file_report in report.file_reports:
        for chapter in file_report.chapters:
            summary = chapter.summary
            for r in chapter.rows:
                rows.append(
                    [
                        file_report.filename,
                        summary.chapter_index,
                        summary.id,
                        summary.title,
                        r.type,
                        r.lemma,
                        r.form,
                        r.parse,
                        r.lemma_occ,
                        r.parse_occ,
                        r.form_occ,
                        r.first_occ_lemma,
                        r.first_occ_parse,
                        r.last_occ_lemma,
                        r.last_occ_parse,
                        r.lemma_first_chapter,
                        r.lemma_last_chapter,
                        r.parse_first_chapter,
                        r.parse_last_chapter,
                    ]
                )
    return rows


def new_lemma_rows(report: MultiFileReport) -> list[list]:
    """One row per lemma, in the chapter that introduces it.

    A lemma is new exactly once in the corpus, so no lemma can appear twice
    here — a repeat would mean two chapters both claiming to be its first, which
    is a fault in the index rather than something to dedupe away.
    """
    rows: list[list] = []
    for file_report in report.file_reports:
        for chapter in file_report.chapters:
            summary = chapter.summary
            # Insertion-ordered: one lemma's parse rows are not adjacent in
            # report order, but the lemma's place in that order is its first
            # appearance in the chapter, which is the order to keep.
            introduced: dict[str, list] = {}
            for r in chapter.rows:
                if r.lemma_first_chapter != summary.chapter_index:
                    continue
                introduced.setdefault(r.lemma, []).append(r)

            for lemma, lemma_rows in introduced.items():
                first = lemma_rows[0]
                forms: list[str] = []
                parses: list[str] = []
                for r in lemma_rows:
                    if r.parse and r.parse not in parses:
                        parses.append(r.parse)
                    # `forms` is the row's full breakdown; `form` alone is only
                    # its most frequent spelling, so a lemma met as `Ὁ` and `ὁ`
                    # under one parse would otherwise lose one of them.
                    for spelling in [f.form for f in r.forms] or [r.form]:
                        if spelling and spelling not in forms:
                            forms.append(spelling)

                rows.append(
                    [
                        file_report.filename,
                        summary.chapter_index,
                        summary.id,
                        summary.title,
                        first.type,
                        lemma,
                        first.lemma_occ,
                        len(forms),
                        JOIN.join(forms),
                        len(parses),
                        JOIN.join(parses),
                        first.lemma_last_chapter,
                        # Chapters from here to the last one carrying it, this
                        # one included: 1 means the word is never seen again.
                        first.lemma_last_chapter - summary.chapter_index + 1,
                    ]
                )
    return rows


def _write_sheet(ws, header: list[str], rows: list[list], widths: dict[str, int]) -> None:
    ws.append(header)
    for row in rows:
        ws.append(row)

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="top")

    # A header that scrolls away is unreadable at this width, and nineteen
    # columns of Greek want a filter more than they want styling.
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"

    for name, width in widths.items():
        ws.column_dimensions[get_column_letter(header.index(name) + 1)].width = width


def _workbook_bytes(rows: list[list], new_rows: list[list]) -> bytes:
    """The whole export as .xlsx bytes."""
    wb = Workbook()
    rows_ws = wb.active
    rows_ws.title = ROW_SHEET
    _write_sheet(
        rows_ws,
        ROW_HEADER,
        rows,
        {"file": 24, "chapter_id": 18, "chapter_title": 24, "lemma": 16, "form": 16, "parse": 20},
    )

    new_ws = wb.create_sheet(NEW_LEMMA_SHEET)
    _write_sheet(
        new_ws,
        NEW_LEMMA_HEADER,
        new_rows,
        {
            "file": 24,
            "chapter_id": 18,
            "chapter_title": 24,
            "lemma": 16,
            "forms": 28,
            "parses": 28,
        },
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(report: MultiFileReport, now: datetime | None = None) -> str:
    """A dated filename naming the source, so a folder of exports stays sortable."""
    stamp = (now or datetime.now()).strftime("%Y%m%d-%H%M")
    names = [re.sub(r"\.docx$", "", f.filename, flags=re.IGNORECASE) for f in report.file_reports]
    # One source file names the export after it; several would make an
    # unreadable filename, so they fall back to a count.
    if len(names) == 1:
        base = re.sub(r"^-+|-+$", "", re.sub(r"[^\w]+", "-", names[0], flags=re.UNICODE))
    else:
        base = f"{len(names)}-files"
    return f"lingua-calc-{base or 'report'}-{stamp}.xlsx"


class Export(NamedTuple):
    """A built workbook and the two figures the UI reports back.

    The counts ride along because the route would otherwise have to walk the
    report a second time to say how much it just exported, and on a corpus-sized
    run that walk is not free.
    """

    filename: str
    content: bytes
    row_count: int
    new_lemma_count: int


def build_export(report: MultiFileReport, now: datetime | None = None) -> Export:
    rows = export_rows(report)
    new_rows = new_lemma_rows(report)
    return Export(
        filename=export_filename(report, now),
        content=_workbook_bytes(rows, new_rows),
        row_count=len(rows),
        new_lemma_count=len(new_rows),
    )
