"""The workbook export (issue #3).

Two sheets with two different grains, so the tests are mostly about what each
one collapses: the flat sheet keeps every chapter × lemma × parse row, and the
new-lemma sheet keeps one row per lemma in the chapter that introduces it.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import load_workbook

from lingua_calc.corpus import CorpusIndex
from lingua_calc.export import (
    NEW_LEMMA_HEADER,
    NEW_LEMMA_SHEET,
    ROW_HEADER,
    ROW_SHEET,
    build_export,
    export_filename,
    export_rows,
    new_lemma_rows,
)
from lingua_calc.models import MultiFileReport
from lingua_calc.pipeline import build_file_reports

from .conftest import GEN, HO, HO_CAP, LOGOI, LOGOS, LOGOU, NOM, VOC, facts_from

# ὁ is met twice in chapter 0 under one parse but two spellings; λόγος is met in
# both chapters under two parses; θεός arrives only in chapter 1. Between them
# that is every case the new-lemma sheet has to get right.
CORPUS = facts_from(
    [
        (HO, HO_CAP, NOM, 0),
        (HO, HO, NOM, 0),
        (LOGOS, LOGOS, NOM, 0),
        (LOGOS, LOGOU, GEN, 0),
        (LOGOS, LOGOS, NOM, 1),
        ("θεός", "θεοῦ", GEN, 1),
    ]
)


def report(facts=None) -> MultiFileReport:
    return MultiFileReport(file_reports=build_file_reports(CorpusIndex(facts or CORPUS)))


def sheet_dicts(rows: list[list], header: list[str]) -> list[dict]:
    return [dict(zip(header, row)) for row in rows]


def new_by_lemma(rep: MultiFileReport) -> dict[str, dict]:
    return {r["lemma"]: r for r in sheet_dicts(new_lemma_rows(rep), NEW_LEMMA_HEADER)}


# -- the flat sheet ---------------------------------------------------------


def test_flat_sheet_keeps_every_chapter_lemma_parse_row():
    rep = report()
    rows = sheet_dicts(export_rows(rep), ROW_HEADER)

    assert len(rows) == sum(len(c.rows) for f in rep.file_reports for c in f.chapters)
    assert [(r["chapter_index"], r["lemma"], r["parse"]) for r in rows] == [
        (0, HO, NOM),
        (0, LOGOS, NOM),
        (0, LOGOS, GEN),
        (1, LOGOS, NOM),
        (1, "θεός", GEN),
    ]


def test_flat_sheet_carries_the_file_and_chapter_the_row_belongs_to():
    row = sheet_dicts(export_rows(report()), ROW_HEADER)[0]

    assert (row["file"], row["chapter_id"], row["chapter_title"]) == (
        "a.docx",
        "1-ch",
        "Chapter 1",
    )


def test_first_occurrence_booleans_agree_with_the_raw_indexes():
    """The booleans are the comparison the raw columns invite; if they can
    disagree, a reader filtering on either one gets a different answer."""
    for row in sheet_dicts(export_rows(report()), ROW_HEADER):
        assert row["first_occ_lemma"] == (row["lemma_first_chapter"] == row["chapter_index"])
        assert row["first_occ_parse"] == (row["parse_first_chapter"] == row["chapter_index"])


# -- the new-lemma sheet ----------------------------------------------------


def test_new_lemmas_land_only_in_the_chapter_that_introduces_them():
    rows = sheet_dicts(new_lemma_rows(report()), NEW_LEMMA_HEADER)

    assert [(r["chapter_index"], r["lemma"]) for r in rows] == [
        (0, HO),
        (0, LOGOS),
        (1, "θεός"),
    ]


def test_a_lemma_never_appears_twice_however_often_it_recurs():
    """λόγος is in both chapters and wears two parses in the first; the sheet is
    one row per lemma, so all of that has to collapse to a single row."""
    rows = [r["lemma"] for r in sheet_dicts(new_lemma_rows(report()), NEW_LEMMA_HEADER)]

    assert rows.count(LOGOS) == 1
    assert len(rows) == len(set(rows))


def test_the_parses_a_new_lemma_wears_are_joined_onto_its_one_row():
    logos = new_by_lemma(report())[LOGOS]

    assert logos["parse_count"] == 2
    assert set(logos["parses"].split(" · ")) == {NOM, GEN}
    assert logos["lemma_occ"] == 2


def test_spellings_come_from_the_full_form_breakdown_not_the_representative():
    """`Ὁ` and `ὁ` are one lemma under one parse — the row's representative
    `form` is only one of them, so counting spellings off it would lose the
    other."""
    ho = new_by_lemma(report())[HO]

    assert ho["form_count"] == 2
    assert set(ho["forms"].split(" · ")) == {HO, HO_CAP}


def test_chapter_span_reaches_the_last_chapter_that_carries_the_lemma():
    rows = new_by_lemma(report())

    # ὁ is introduced and dropped in chapter 0; λόγος runs on into chapter 1.
    assert (rows[HO]["lemma_last_chapter"], rows[HO]["chapter_span"]) == (0, 1)
    assert (rows[LOGOS]["lemma_last_chapter"], rows[LOGOS]["chapter_span"]) == (1, 2)


def test_a_lemma_introduced_by_a_second_file_is_new_there():
    """Chapter indexes are corpus-wide, so "new" spans files: a word first met
    in the second file is new in the second file, not new twice."""
    facts = facts_from([(LOGOS, LOGOS, NOM, 0)], filename="a.docx") + facts_from(
        [(LOGOS, LOGOI, VOC, 1), ("θεός", "θεός", NOM, 1)], filename="b.docx"
    )
    rows = sheet_dicts(new_lemma_rows(report(facts)), NEW_LEMMA_HEADER)

    assert [(r["file"], r["lemma"]) for r in rows] == [("a.docx", LOGOS), ("b.docx", "θεός")]


# -- the workbook itself ----------------------------------------------------


def test_workbook_has_both_sheets_with_their_headers():
    export = build_export(report())
    wb = load_workbook(io.BytesIO(export.content))

    assert wb.sheetnames == [ROW_SHEET, NEW_LEMMA_SHEET]
    assert [c.value for c in wb[ROW_SHEET][1]] == ROW_HEADER
    assert [c.value for c in wb[NEW_LEMMA_SHEET][1]] == NEW_LEMMA_HEADER


def test_workbook_rows_match_the_grains_and_the_counts_reported_with_them():
    rep = report()
    export = build_export(rep)
    wb = load_workbook(io.BytesIO(export.content))

    assert export.row_count == len(export_rows(rep))
    assert export.new_lemma_count == len(new_lemma_rows(rep))
    # +1 for the header row on each sheet.
    assert wb[ROW_SHEET].max_row == export.row_count + 1
    assert wb[NEW_LEMMA_SHEET].max_row == export.new_lemma_count + 1


def test_greek_survives_the_round_trip():
    """The whole point of leaving CSV behind: no BOM, no codepage, no mojibake."""
    wb = load_workbook(io.BytesIO(build_export(report()).content))
    lemmas = {row[5].value for row in wb[ROW_SHEET].iter_rows(min_row=2)}

    assert {HO, LOGOS, "θεός"} <= lemmas


def test_header_is_frozen_so_it_survives_scrolling():
    wb = load_workbook(io.BytesIO(build_export(report()).content))

    assert wb[ROW_SHEET].freeze_panes == "A2"
    assert wb[NEW_LEMMA_SHEET].freeze_panes == "A2"


# -- filename ---------------------------------------------------------------


def when() -> datetime:
    return datetime(2026, 8, 13, 9, 5)


def test_one_source_file_names_the_export():
    rep = report()

    assert export_filename(rep, when()) == "lingua-calc-a-20260813-0905.xlsx"


def test_several_source_files_fall_back_to_a_count():
    facts = facts_from([(LOGOS, LOGOS, NOM, 0)], filename="a.docx") + facts_from(
        [("θεός", "θεός", NOM, 1)], filename="b.docx"
    )

    assert export_filename(report(facts), when()) == "lingua-calc-2-files-20260813-0905.xlsx"
